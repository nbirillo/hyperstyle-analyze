import logging.config
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import load_workbook, Workbook

logger = logging.getLogger(__name__)


def remove_sheet(path: Union[str, Path], sheet_name: str, to_raise_error: bool = False) -> None:
    try:
        workbook = load_workbook(path)
        workbook.remove(workbook[sheet_name])
        workbook.save(path)

    except KeyError as e:
        message = f'Sheet with specified name: {sheet_name} does not exist.'
        if to_raise_error:
            logger.exception(message)
            raise e
        else:
            logger.info(message)


def create_workbook(path: Union[str, Path]) -> Workbook:
    workbook = Workbook()
    workbook.save(path)
    return workbook


def write_df_to_xlsx(df: pd.DataFrame,
                     path: Union[str, Path],
                     sheet_name: str = 'Sheet1',
                     mode: str = 'w',
                     index: bool = False,
                     header: bool = True) -> None:
    with pd.ExcelWriter(path, mode=mode) as writer:
        start_row = writer.sheets[sheet_name].max_row if mode == 'a' else 0
        df.to_excel(writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=index,
                    header=header)


def read_df_from_xlsx(path: Union[str, Path], sheet_name: str = 'Sheet1') -> pd.DataFrame:
    return pd.read_excel(path, sheet_name)
